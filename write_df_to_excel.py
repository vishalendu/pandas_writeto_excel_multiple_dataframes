## REF: https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas/47740262#47740262

import pandas as pd
from openpyxl import load_workbook,Workbook
import os

START_COL=7
START_ROW=6

# Open existing file or create a new one
def getExcel(PATH):
    if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
        print("File exists and is readable. Reading existing file")
        book = load_workbook(PATH)
    else:
        print("Either the file is missing or not readable. Creating new file.")
        book = Workbook()
        ws = book.active
        ws.title = "Sheet1"
        book.save(filename=PATH)
    return book

def write_df_col(PATH,sheet,df,startc,indexWrite):
    try:
        book = getExcel(PATH)  
        writer = pd.ExcelWriter(PATH, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        print(f"writer.sheets = {writer.sheets}")
        df.to_excel(writer, sheet, startcol=startc,startrow=START_ROW,index=indexWrite)
        writer.save()
    except:
        book.close()
    finally:
        book.close()


if __name__ == "__main__":        
    test1df = pd.DataFrame({'Data': [3, 14, 25, 36]})
    test2df = pd.DataFrame({'Data': [13, 24, 35, 46]})
    test3df = pd.DataFrame({'Data': [23, 34, 45, 56]})

    # testdict= {1:test1df,2:test2df,3:test3df}
    
    #STEP1 -- Update 2 values (Check Excel should have 2 Data columns)
    #testdict= {1:test1df,2:test2df}
    #STEP2 -- Add one more columns
    testdict= {3:test3df}

    File='./sample.xlsx'

    listcol=[0,2,3]
    for key, df in testdict.items():
        if(key==1):
            write_df_col(File,"Sheet1",df,START_COL,True)
        else:
            write_df_col(File,"Sheet1",df,START_COL+ listcol[key-1],False)